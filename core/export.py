import csv
import json
import os
from typing import List
from openpyxl import Workbook, load_workbook
from core.base_parser import VehicleListing
from dataclasses import asdict

def _get_headers() -> List[str]:
    """Получить список заголовков из дата-класса"""
    return list(VehicleListing.__dataclass_fields__.keys())

def export_to_csv(listings: List[VehicleListing], filepath: str):
    """Экспорт списка объявлений в CSV файл. Добавляет данные, если файл существует."""
    file_exists = os.path.isfile(filepath)
    headers = _get_headers()
    
    with open(filepath, 'a', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=headers)
        
        if not file_exists:
            writer.writeheader()
            
        for listing in listings:
            data = asdict(listing)
            data['parsed_at'] = data['parsed_at'].isoformat()
            writer.writerow(data)

def export_to_json(listings: List[VehicleListing], filepath: str):
    """Экспорт списка объявлений в JSON файл."""
    data = []
    
    # Чтение существующих данных
    if os.path.isfile(filepath):
        try:
            with open(filepath, 'r', encoding='utf-8') as jsonfile:
                data = json.load(jsonfile)
        except Exception:
            pass
            
    # Добавление новых
    for listing in listings:
        dict_data = asdict(listing)
        dict_data['parsed_at'] = dict_data['parsed_at'].isoformat()
        data.append(dict_data)
        
    with open(filepath, 'w', encoding='utf-8') as jsonfile:
        json.dump(data, jsonfile, ensure_ascii=False, indent=4)

def export_to_excel(listings: List[VehicleListing], filepath: str):
    """Экспорт списка объявлений в Excel с использованием openpyxl."""
    headers = _get_headers()
    
    if os.path.isfile(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        # Форматирование заголовков можно добавить здесь
        
    for listing in listings:
        data = asdict(listing)
        data['parsed_at'] = data['parsed_at'].replace(tzinfo=None) # Excel не поддерживает tz
        row = [data[h] for h in headers]
        ws.append(row)
        
    wb.save(filepath)
